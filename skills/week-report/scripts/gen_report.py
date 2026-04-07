#!/usr/bin/env python3
"""
gen_report.py — 前端工作周报 Excel 生成脚本
用法：python3 gen_report.py --template <模板路径> --output <输出路径> --name <姓名> --focus <工作重点> --items <条目JSON>

条目 JSON 格式：
[
  {"work": "工作内容描述", "module": "所属模块"},
  ...
]
"""
import argparse
import json
import shutil
import sys

def check_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        import subprocess
        print("[week-report] 正在安装 openpyxl...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
        import openpyxl
        return openpyxl

def calc_week_of_month(date_str):
    """计算指定日期是当月第几周（包含该周一的自然周）"""
    from datetime import datetime, timedelta
    d = datetime.strptime(date_str, "%Y-%m-%d")
    first_day = d.replace(day=1)
    # 当月1号是周几（0=周一）
    first_weekday = first_day.weekday()
    # 当月1号所在周的周一
    first_monday = first_day - timedelta(days=first_weekday)
    # 目标日期所在周的周一
    target_monday = d - timedelta(days=d.weekday())
    week_num = (target_monday - first_monday).days // 7 + 1
    return week_num

def generate(template_path, output_path, name, focus, items):
    openpyxl = check_openpyxl()
    from openpyxl import load_workbook

    # 复制模板，不修改原文件
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # 取消数据区所有合并（保留标题行 A1:E1）
    merges_to_remove = [str(m) for m in ws.merged_cells.ranges if str(m) != "A1:E1"]
    for m in merges_to_remove:
        ws.unmerge_cells(m)

    # 写入数据行（从第3行开始）
    data_start = 3
    for i, item in enumerate(items):
        row = data_start + i
        ws.cell(row=row, column=1).value = item["work"]
        ws.cell(row=row, column=2).value = item["module"]
        ws.cell(row=row, column=3).value = name if i == 0 else ""
        ws.cell(row=row, column=4).value = focus if i == 0 else ""
        ws.cell(row=row, column=5).value = ""

    # 清空多余行内容
    data_end = data_start + len(items) - 1
    for row in range(data_end + 1, ws.max_row + 1):
        for col in range(1, 6):
            ws.cell(row=row, column=col).value = None

    # 动态合并 C、D 列（人员 & 工作重点）
    if len(items) > 1:
        ws.merge_cells(f"C{data_start}:C{data_end}")
        ws.merge_cells(f"D{data_start}:D{data_end}")

    wb.save(output_path)
    print(f"[week-report] 周报已生成：{output_path}")

def resolve_versioned_path(output_path):
    """
    自动版本化输出路径，避免覆盖已有文件。
    规则：
      - 不存在 → 直接使用（如 xxx.xlsx）
      - 已存在 → 查找下一个可用版本号
        xxx.xlsx 存在 → xxx_v2.xlsx → xxx_v3.xlsx ...
    """
    import os
    if not os.path.exists(output_path):
        return output_path

    base, ext = os.path.splitext(output_path)
    version = 2
    while True:
        candidate = f"{base}_v{version}{ext}"
        if not os.path.exists(candidate):
            return candidate
        version += 1


def main():
    parser = argparse.ArgumentParser(description="生成前端工作周报 Excel")
    parser.add_argument("--template", required=True, help="模板 Excel 路径")
    parser.add_argument("--output",   required=True, help="输出 Excel 路径（同周重复生成时自动追加 _v2/_v3）")
    parser.add_argument("--name",     required=True, help="姓名")
    parser.add_argument("--focus",    required=True, help="工作重点（全局一句话总结）")
    parser.add_argument("--items",    required=True, help="工作条目 JSON 字符串")
    args = parser.parse_args()

    items = json.loads(args.items)
    if not items:
        print("[week-report] 错误：工作条目为空", file=sys.stderr)
        sys.exit(1)

    output_path = resolve_versioned_path(args.output)
    generate(args.template, output_path, args.name, args.focus, items)

if __name__ == "__main__":
    main()
